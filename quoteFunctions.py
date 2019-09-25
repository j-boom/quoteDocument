#Column Widths
from docx.shared import Cm, Inches
import openpyxl, datetime, os

def set_col_widths(table, height, row1, row2, row3, row4):
    widths = [Inches(row1), Inches(row2), Inches(row3), Inches(row4)]
    for row in table.rows:
        row.height = Inches(height)
        for idx, width in enumerate(widths):
            row.cells[idx].width = width

def getScopeList(excelFile):
    j = 7
    wb = openpyxl.load_workbook(excelFile)
    sheet = wb["Sheet1"]
    scopeList = []

    while sheet.cell(j,2).value != None:
        scopeList.append(sheet.cell(j,2).value)
        j += 1
    scopeList.append(j)
    return scopeList

def getClarifyList(gslIDX, excelFile):
    j= gslIDX + 1
    wb = openpyxl.load_workbook(excelFile)
    sheet = wb["Sheet1"]
    clarifyList = []

    while sheet.cell(j,2).value != None:
        clarifyList.append(sheet.cell(j,2).value)
        j += 1
    return clarifyList

def updateTracker(quoteID, quoteTitle, companyName, contact, cost):
    #define variables
    rowNum = 3
    allQuotes = []
    
    #open the right tracker
    xlInputPath = os.path.join("C:\\users\\jbergrj\\OneDrive - Johnson Controls\\Quotes")
    os.chdir(xlInputPath)
    wb = openpyxl.load_workbook("1. tracker.xlsx")
    sheet = wb["L&M Tracker"]

    #validate that this quote is not already in the tracker
    while sheet.cell(rowNum,2).value != None:
        allQuotes.append(sheet.cell(rowNum,2).value)
        rowNum += 1
    print (allQuotes)

    if quoteID in allQuotes:
        print("This quote already exists")
        return

    else: 
        rowNum = 3
        while sheet.cell(rowNum,2).value != None:
        # count through column B to find the next empty row
            rowNum += 1
        
    #begin data entry
  
    sheet.cell(rowNum,1).value = "L&M"
    sheet.cell(rowNum,2).value = quoteID
    sheet.cell(rowNum,3).value = quoteTitle
    sheet.cell(rowNum,4).value = datetime.date.today().strftime("%m/%d/%Y")
    sheet.cell(rowNum,5).value = companyName
    sheet.cell(rowNum,6).value = contact
    sheet.cell(rowNum,9).value = cost

    wb.save("tracker.xlsx")
