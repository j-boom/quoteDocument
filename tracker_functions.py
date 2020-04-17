import datetime, os

import openpyxl
from docx import Document

class Tracker:
    def __init__(self):
        self.tracker_path = os.path.join("C:\\Users\\jbergren\\OneDrive - ENGINEERED SERVICES INC\\Quotes\\Quote Numbers")

    def create_quote_number(self, quote_number, previous):
        
        os.chdir(self.tracker_path)
        wb = openpyxl.load_workbook("1. tracker.xlsx")
        ws = wb['Proposal Tracker']

        cell_row = ""

        if previous:
            for row in ws.iter_rows(min_row=1, max_col=1, max_row=len(ws['A'])):
                for cell in row:
                    if cell.value == quote_number:
                        cell_row = cell.row
                        break
            new_float = float(quote_number[2:]) + 0.01
            quote_number = 'JB{}'.format(new_float)
        else:
            last_cell = len(ws['A'])
            last_quote_number = int(ws["A{}".format(last_cell)].value[4:])
            quote_number = "JB{}{}".format(datetime.datetime.today().strftime('%y'), last_quote_number + 1)

        return quote_number, cell_row

    def update(self, quote, previous, tracker_row):    
        os.chdir(self.tracker_path)
        wb = openpyxl.load_workbook("1. tracker.xlsx")
        ws = wb['Proposal Tracker']

        if not previous:
            tracker_row = len(ws['A']) + 1

        ws.cell(tracker_row,1).value = quote.quote_number
        ws.cell(tracker_row,2).value = quote.project_name
        ws.cell(tracker_row,3).value = datetime.date.today().strftime("%m/%d/%Y")
        ws.cell(tracker_row,4).value = quote.project_site
        ws.cell(tracker_row,5).value = quote.company_name
        ws.cell(tracker_row,6).value = quote.point_of_contact
        ws.cell(tracker_row,8).value = quote.price

        wb.save("1. tracker.xlsx")