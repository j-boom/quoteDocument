import datetime, os

from docx.shared import Cm, Inches
from openpyxl.styles import Font
import openpyxl

def set_col_widths(table, height, col1, col2):
    widths = [Inches(col1), Inches(col2)]
    for row in table.rows:
        row.height = Inches(height)
        for idx, width in enumerate(widths):
            row.cells[idx].width = width

def get_list(excel_file, start_index):
    j= start_index + 10
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb["Sheet1"]
    collected_list = []


    while sheet['B{}'.format(j)].value != None:
        collected_list.append(sheet['B{}'.format(j)].value)
        j += 1
    return collected_list

def clean_up_input(previous):
    if not previous:
        excel_input_path = os.path.join("C:\\Users\\jbergren\\OneDrive - ENGINEERED SERVICES INC\\Quotes\\Quote Numbers")
        os.chdir(excel_input_path)
        wb = openpyxl.load_workbook("quote_input.xlsx")
        ws = wb['Sheet1']

        wb.remove(ws)
        
        wb.create_sheet(title='Sheet1', index=0)
        ws = wb['Sheet1']

        form_cells = ['A{}'.format(i) for i in range(1, 15)]
        form = [
            'Quote ID:',
            'Point of Contact:',
            'Company:',
            'Address 1:',
            'Address 2:',
            'Project Site:',
            'Project Name:',
            'Project Price (Must be a text type):',
            'The cell next to this one must stay blank.',
            'Scope of work:',
            'The cell next to this one must stay blank.  Insert cells above it to add to the scope of work.',
            'Exclusions:',
            'The cell next to this one must stay blank.  Insert cells above it to add exclusions.',
            'Clarifications:',
        ]
        form_cells.append('B1')
        form.append('New Quote')
        form_dictionary = dict(zip(form_cells, form))

        instructions_cells = ['C4', 'C5', 'C8', 'C10', 'C12', 'C14']

        instructions = [
        'ex: 123 Melody Lane',
        'ex: Sterling, VA 12345',
        'Cell B8 must be type text',
        '<---- Start Scope of work in this cell',
        '<---- Start Exclusions list in this cell',
        '<---- Start Clarifications list in this cell',
        ]
        instructions_dictionary = dict(zip(instructions_cells, instructions))

        for row in ws.iter_rows(min_row=1, max_col=3, max_row=14):
            for cell in row:
                if cell.coordinate in instructions_cells:
                    cell.value = instructions_dictionary.get(cell.coordinate)
                if cell.coordinate in form_cells:
                    cell.value = form_dictionary.get(cell.coordinate)
                if cell.coordinate in form_cells and form_dictionary.get(cell.coordinate).startswith('The cell next'):
                    cell.font = Font(name='Segoe UI', size=10, bold=False, italic=True, color="003366")
                elif cell.coordinate in form_cells or cell.coordinate in instructions_cells:
                    cell.font = Font(name='Calibri Light', size=12, bold=True, italic=False)
                else:
                    cell.font = Font(name='Calibri LIght', size=12, bold=False, italic=False)
                i += 1
        
        ws.column_dimensions['A'].width = 21
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 25

        wb.save("quote_input.xlsx")