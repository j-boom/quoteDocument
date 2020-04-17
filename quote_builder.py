import datetime, os

import openpyxl

from make_quote_document import make_quote, Quote
from terms_conditions import TandCs
from tracker_functions import Tracker
from quote_input_functions import get_list, clean_up_input

def main():
    rep_name = 'Jim Bergren'
    rep_email = 'jbergren@engineeredservices.com'
    excel_path ="C:\\Users\\jbergren\\OneDrive - ENGINEERED SERVICES INC\\Quotes\\Quote Numbers"

    #Get User Input
    user_file_name = raw_input("Is this a new project?  Enter (y)es or (n)o: ")

    if user_file_name.startswith("y") or user_file_name.startswith("Y"):
        user_file_name = "quote_input.xlsx"
        previous = False
    else:
        user_file_name = raw_input("Enter the filename you wish to use (don't forget the extension): ")
        previous = True
    print ("Getting data from {}\\{}".format(excel_path, user_file_name))
    
    #instantiate an instance of Quote
    quote = Quote()
    tracker = Tracker()

    excel_file = "{}\\{}".format(excel_path, user_file_name)

    data = get_list(excel_file, -9)
    scope_list = get_list(excel_file, 0)
    exclusion_list = get_list(excel_file, len(scope_list) + 1)
    clarification_list = get_list(excel_file, len(scope_list) + len(exclusion_list) + 2)
    quote.quote_number, tracker_row = tracker.create_quote_number(data[0], previous)
    quote.point_of_contact = data[1]
    quote.company_name = data[2]
    quote.company_address_one = data[3]
    quote.company_address_two = data[4]
    quote.project_site = data[5]
    quote.project_name = data[6]
    quote.price = data[7]

    make_quote(excel_file, scope_list, exclusion_list, clarification_list, rep_name, rep_email, TandCs, quote)
    
    excel_save_name = "{}-{}-{}-{}.xlsx".format(quote.quote_number, quote.project_site, quote.project_name, quote.time_stamp)
    print('...Saving inputs to {}'.format(excel_save_name))
    path = os.path.join(excel_path)
    os.chdir(path)
    wb = openpyxl.load_workbook(user_file_name)
    ws = wb["Sheet1"]
    ws["B1"].value = quote.quote_number
    ws["C1"].value = "File Name:"
    ws["D1"].value = excel_save_name
    wb.save(excel_save_name)

    tracker.update(quote, previous, tracker_row)

# ###-----------------------------------------------------------------CLEAN UP INPUT FOR NEXT QUOTE-------------------------------------------------------------###
    print('...Cleaning up input file')
    clean_up_input(previous)
        

# ###-----------------------------------------------------------------CHECK SUCCESS-------------------------------------------------------------###
#             new_proposal = os.path.join(f"C:\\Users\\jbergren\\OneDrive - ENGINEERED SERVICES INC\\Quotes\\{output_file}")
#         try:
#             if os.path.isfile(new_proposal):
#                 print('Great!  Go win this job!')
#         except:
#             print(f'Uh oh... I couldn\'t find the file I just saved')

if __name__ == "__main__":
    main()